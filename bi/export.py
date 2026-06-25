"""
BI export (Phase 5).

Produces a clean, BI-ready data source from AlphaForge's own analytics:
  bi/output/alphaforge_bi.xlsx   (multi-sheet, formatted)
  bi/output/*.csv                (one per table — easiest to connect in Tableau)

Connect Tableau or Power BI to these files and build the views described in
bi/DASHBOARD_SPEC.md. Re-run after a data refresh to update the dashboard source.

Run:  python -m bi.export
"""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from data_pipeline.db import connect
from finance.factors import factor_table
from finance.valuation import dcf
from models.predict import predict_latest
from risk.montecarlo_var import monte_carlo_var

OUT = Path(__file__).resolve().parent / "output"
HEADER_FILL = PatternFill("solid", start_color="1F2A44")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial")


def _gather():
    rows = factor_table()
    holdings = []
    for r in rows:
        t = r["ticker"]
        try:
            d = dcf(t)
            fair, up = d.fair_value_per_share, d.upside_pct
        except Exception:
            fair, up = None, None
        try:
            ml = predict_latest(t).get("p_outperform_sector_1m")
        except Exception:
            ml = None
        holdings.append({
            "ticker": t, "sector": r["sector"], "price": r["price"],
            "market_cap_mm": r["market_cap"], "pe": r["pe"], "ev_ebitda": r["ev_ebitda"],
            "roic_pct": r["roic_pct"], "ebitda_margin_pct": r["ebitda_margin_pct"],
            "net_debt_ebitda": r["net_debt_ebitda"],
            "dcf_fair_value": fair, "dcf_upside_pct": up, "ml_p_outperform": ml,
        })

    # sector exposure
    sectors = {}
    for h in holdings:
        s = sectors.setdefault(h["sector"], {"count": 0, "mcap": 0.0, "roic": [], "evebitda": []})
        s["count"] += 1
        s["mcap"] += h["market_cap_mm"] or 0
        if h["roic_pct"] is not None:
            s["roic"].append(h["roic_pct"])
        if h["ev_ebitda"] is not None:
            s["evebitda"].append(h["ev_ebitda"])
    sector_rows = [{
        "sector": k, "n_names": v["count"], "market_cap_mm": round(v["mcap"], 0),
        "avg_roic_pct": round(sum(v["roic"]) / len(v["roic"]), 1) if v["roic"] else None,
        "avg_ev_ebitda": round(sum(v["evebitda"]) / len(v["evebitda"]), 1) if v["evebitda"] else None,
    } for k, v in sorted(sectors.items())]

    # risk: standalone VaR per name + equal-weight portfolio
    risk_rows = []
    for h in holdings:
        try:
            rv = monte_carlo_var([h["ticker"]], n_paths=40_000)
            risk_rows.append({"ticker": h["ticker"], "sector": h["sector"],
                              "var_95_pct": rv["VaR_pct"], "cvar_95_pct": rv["CVaR_pct"]})
        except Exception:
            pass
    port = monte_carlo_var([h["ticker"] for h in holdings], n_paths=60_000)
    risk_rows.append({"ticker": "PORTFOLIO(EW)", "sector": "ALL",
                      "var_95_pct": port["VaR_pct"], "cvar_95_pct": port["CVaR_pct"]})
    return holdings, sector_rows, risk_rows


def _write_sheet(ws, title, rows):
    ws.title = title
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([r[h] for h in headers])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 2)
    ws.freeze_panes = "A2"


def export():
    OUT.mkdir(parents=True, exist_ok=True)
    holdings, sector_rows, risk_rows = _gather()

    wb = Workbook()
    _write_sheet(wb.active, "Holdings", holdings)
    _write_sheet(wb.create_sheet(), "SectorExposure", sector_rows)
    _write_sheet(wb.create_sheet(), "Risk", risk_rows)

    # Summary sheet with live formulas referencing Holdings
    s = wb.create_sheet("Summary")
    n = len(holdings) + 1
    s["A1"] = "Metric"; s["B1"] = "Value"
    for c in ("A1", "B1"):
        s[c].fill = HEADER_FILL; s[c].font = HEADER_FONT
    s["A2"] = "Names covered";       s["B2"] = f"=COUNTA(Holdings!A2:A{n})"
    s["A3"] = "Total market cap ($mm)"; s["B3"] = f"=SUM(Holdings!D2:D{n})"
    s["A4"] = "Avg DCF upside (%)";  s["B4"] = f"=AVERAGE(Holdings!K2:K{n})"
    s["A5"] = "Avg ROIC (%)";        s["B5"] = f"=AVERAGE(Holdings!G2:G{n})"
    s["A6"] = "Median EV/EBITDA (x)"; s["B6"] = f"=MEDIAN(Holdings!F2:F{n})"
    for r in range(2, 7):
        s[f"A{r}"].font = BODY_FONT; s[f"B{r}"].font = BODY_FONT
    s.column_dimensions["A"].width = 26
    s.column_dimensions["B"].width = 18
    wb.move_sheet("Summary", -(len(wb.sheetnames) - 1))  # put Summary first

    xlsx_path = OUT / "alphaforge_bi.xlsx"
    wb.save(xlsx_path)

    # CSVs (Tableau-friendly)
    for name, rows in [("holdings", holdings), ("sector_exposure", sector_rows), ("risk", risk_rows)]:
        if rows:
            with open(OUT / f"{name}.csv", "w", newline="") as f:
                w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
                w.writeheader(); w.writerows(rows)

    print(f"Wrote {xlsx_path} and CSVs to {OUT}/")
    return xlsx_path


if __name__ == "__main__":
    from data_pipeline.sample_data import generate
    with connect() as conn:
        from data_pipeline.db import init_db
        init_db()
        n = conn.execute("SELECT COUNT(*) c FROM prices").fetchone()["c"]
    if not n:
        generate()
    export()
